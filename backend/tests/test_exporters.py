"""
Tests for the offline PDF / Excel exporters (no external API calls).
"""

import io
from datetime import date

from openpyxl import load_workbook

from app import engine, exporters, seed

EVAL = date(2026, 4, 15)


def _analyzed():
    return [engine.analyze_record(r, EVAL) for r in seed.generate_records()]


def test_pdf_bytes_are_a_valid_pdf():
    pdf = exporters.report_pdf_bytes(_analyzed(), EVAL)
    assert isinstance(pdf, bytes) and len(pdf) > 1000
    assert pdf[:5] == b"%PDF-"  # PDF magic header


def test_xlsx_has_exceptions_and_summary_sheets():
    data = _analyzed()
    xlsx = exporters.analyzed_xlsx_bytes(data, EVAL)
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["Exceptions", "Summary"]

    ws = wb["Exceptions"]
    # One row per analyzed record (plus the header row).
    assert ws.max_row == len(data) + 1
    header = [c.value for c in ws[1]]
    for col in ("Computed risk", "Alerts", "Recommendation"):
        assert col in header

    summary = wb["Summary"]
    assert summary["A1"].value == "Exception Portfolio Summary"


def test_xlsx_exc_00145_row_is_critical():
    data = _analyzed()
    xlsx = exporters.analyzed_xlsx_bytes(data, EVAL)
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Exceptions"]
    rows = {r[0].value: r for r in ws.iter_rows(min_row=2)}
    rec = rows["EXC-00145"]
    cells = {c.column_letter: c.value for c in rec}
    # Computed-risk column should read CRITICAL for the acceptance case.
    assert "CRITICAL" in [v for v in cells.values() if isinstance(v, str)]
