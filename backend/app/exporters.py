"""
exporters.py — offline PDF / Excel renderers for the portfolio report.

Kept separate from the pure `report.py` so the heavy rendering libraries
(reportlab, openpyxl) stay isolated and `report.py`/`engine.py` remain
dependency-light and trivially unit-testable. Both functions return bytes and
make no file or network I/O — 100% offline (spec §1/§10.4).
"""

from __future__ import annotations

import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app import engine, report

# Shared palette for both renderers (mirrors the UI "security console" accent).
_INK = colors.HexColor("#0f172a")
_ACCENT = colors.HexColor("#0d9488")
_MUTED = colors.HexColor("#64748b")
_RISK_FILL = {
    "CRITICAL": "FFE4E6",
    "HIGH": "FFEDD5",
    "MEDIUM": "FEF9C3",
    "LOW": "DCFCE7",
}

TYPE_LABELS = report.TYPE_LABELS


# ---------------------------------------------------------------------------
# PDF — a clean, management-ready rendering of the §5 portfolio report.
# ---------------------------------------------------------------------------


def report_pdf_bytes(
    records: list[dict], eval_date: date = engine.DEFAULT_EVALUATION_DATE
) -> bytes:
    """Render the portfolio/audit report as a formatted PDF (bytes)."""
    d = report.build_report_data(records, eval_date)
    window_start = eval_date - timedelta(days=90)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title="Exception Portfolio Summary",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    base = getSampleStyleSheet()
    h_title = ParagraphStyle(
        "title", parent=base["Title"], textColor=_INK, fontSize=20, spaceAfter=2
    )
    h_sub = ParagraphStyle(
        "sub", parent=base["Normal"], textColor=_MUTED, fontSize=9, spaceAfter=10
    )
    h_section = ParagraphStyle(
        "section",
        parent=base["Heading2"],
        textColor=_ACCENT,
        fontSize=12,
        spaceBefore=12,
        spaceAfter=6,
        alignment=TA_LEFT,
    )
    body = ParagraphStyle("body", parent=base["Normal"], fontSize=9.5, leading=14)

    story: list = []
    story.append(Paragraph("Exception Portfolio Summary", h_title))
    story.append(
        Paragraph(
            f"Report date {eval_date.isoformat()} &nbsp;·&nbsp; "
            f"Last 90 days ({window_start.isoformat()} to {eval_date.isoformat()})"
            " &nbsp;·&nbsp; Sunset GRC · Approach Option A",
            h_sub,
        )
    )

    # --- Executive summary --------------------------------------------------
    story.append(Paragraph("Executive summary", h_section))
    exec_rows = [
        ["Total active exceptions", str(d["total_active"])],
        ["  HIGH risk (immediate attention)", str(d["high"])],
        ["  MEDIUM risk", str(d["medium"])],
        ["  LOW risk", str(d["low"])],
        [
            "Expiring this month",
            f"{d['expiring_this_month']} ({d['expiring_due']} due for renewal decision)",
        ],
        ["Expired (not revoked) — should be closed", str(d["expired_not_revoked"])],
    ]
    story.append(_kv_table(exec_rows))

    # --- Breakdown by type --------------------------------------------------
    story.append(Paragraph("Breakdown by type (active)", h_section))
    type_rows = [
        ["Admin / Root access", str(d["by_type"]["admin_access"]), "HIGH RISK"],
        ["Firewall rules", str(d["by_type"]["firewall_rule_open"]), "MEDIUM RISK"],
        ["Encryption waivers", str(d["by_type"]["encryption_waiver"]), "HIGH RISK"],
        ["Other (data / dev)", str(d["other"]), "LOW/MEDIUM RISK"],
    ]
    story.append(
        _grid_table(
            ["Type", "Count", "Inherent risk"],
            type_rows,
            col_widths=[80 * mm, 25 * mm, 45 * mm],
        )
    )

    # --- Top high-risk exceptions ------------------------------------------
    story.append(Paragraph("Top high-risk exceptions", h_section))
    top_rows = []
    for i, r in enumerate(d["top"], start=1):
        top_rows.append(
            [
                str(i),
                r.get("requester", "?"),
                TYPE_LABELS.get(r["type"], r["type"]),
                r["computed_risk_level"],
                r.get("start_date") or "—",
                report._primary_flag(r),
            ]
        )
    top_table = _grid_table(
        ["#", "Requester", "Type", "Risk", "Since", "Flag"],
        top_rows,
        col_widths=[8 * mm, 28 * mm, 32 * mm, 20 * mm, 24 * mm, 38 * mm],
        risk_col=3,
    )
    story.append(top_table)

    # --- Recommendations ----------------------------------------------------
    story.append(Paragraph("Recommendations", h_section))
    for r in d["top"]:
        story.append(
            Paragraph(f"→ {r.get('recommendation', 'Review')}", body)
        )

    # --- Audit readiness ----------------------------------------------------
    story.append(Paragraph("Next audit readiness", h_section))
    audit_rows = [
        ["All exceptions documented", "Yes"],
        ["Approvals recorded", f"{d['pct_approved']}%"],
        ["Exceptions overdue for review", str(d["overdue_for_review"])],
        ["Not revoked after expiry", str(d["not_revoked_after_expiry"])],
    ]
    story.append(_kv_table(audit_rows))
    story.append(Spacer(1, 8 * mm))
    story.append(
        Paragraph(
            "Generated offline by the Sunset risk engine — evaluation date is "
            "configurable, never the system clock.",
            h_sub,
        )
    )

    doc.build(story)
    return buf.getvalue()


def _kv_table(rows: list[list[str]]) -> Table:
    t = Table(rows, colWidths=[110 * mm, 64 * mm], hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("TEXTCOLOR", (0, 0), (0, -1), _INK),
                ("TEXTCOLOR", (1, 0), (1, -1), _ACCENT),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ]
        )
    )
    return t


def _grid_table(
    header: list[str],
    rows: list[list[str]],
    col_widths: list[float],
    risk_col: int | None = None,
) -> Table:
    data = [header] + rows
    t = Table(data, colWidths=col_widths, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), _INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if risk_col is not None:
        for i, row in enumerate(rows, start=1):
            fill = _RISK_FILL.get(row[risk_col])
            if fill:
                style.append(
                    ("BACKGROUND", (risk_col, i), (risk_col, i), colors.HexColor("#" + fill))
                )
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------------------
# Excel — analyzed records sheet + a summary sheet.
# ---------------------------------------------------------------------------

_XLSX_COLUMNS = [
    ("exception_id", "Exception ID"),
    ("type", "Type"),
    ("requester", "Requester"),
    ("approver", "Approver"),
    ("justification", "Justification"),
    ("start_date", "Start date"),
    ("end_date", "End date (expiry)"),
    ("status", "Status"),
    ("risk_level", "Input risk"),
    ("renewal_count", "Renewals"),
    ("computed_risk_level", "Computed risk"),
    ("alerts", "Alerts"),
    ("recommendation", "Recommendation"),
    ("framework_tags", "Framework tags"),
    ("cia_tags", "CIA tags"),
    ("days_past_expiry", "Days past expiry"),
]


def analyzed_xlsx_bytes(
    records: list[dict], eval_date: date = engine.DEFAULT_EVALUATION_DATE
) -> bytes:
    """Workbook with an analyzed-records sheet plus a summary sheet (bytes)."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    wrap = Alignment(vertical="top", wrap_text=True)

    # --- Sheet 1: Exceptions (one row per record) ---------------------------
    ws = wb.active
    ws.title = "Exceptions"
    keys = [k for k, _ in _XLSX_COLUMNS]
    ws.append([label for _, label in _XLSX_COLUMNS])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r in records:
        row = []
        for k in keys:
            v = r.get(k, "")
            if isinstance(v, list):
                v = " | ".join(v)
            row.append(v)
        ws.append(row)

    # Colour the computed-risk column and wrap long text.
    risk_idx = keys.index("computed_risk_level") + 1
    for i in range(2, ws.max_row + 1):
        cell = ws.cell(row=i, column=risk_idx)
        fill = _RISK_FILL.get(str(cell.value))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True)
        ws.cell(row=i, column=keys.index("alerts") + 1).alignment = wrap
        ws.cell(row=i, column=keys.index("justification") + 1).alignment = wrap

    _autosize(ws, max_width=48)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{ws.max_row}"

    # --- Sheet 2: Summary ---------------------------------------------------
    d = report.build_report_data(records, eval_date)
    s = wb.create_sheet("Summary")
    s.append(["Exception Portfolio Summary"])
    s["A1"].font = Font(bold=True, size=14)
    s.append(["Report date", eval_date.isoformat()])
    s.append([])
    summary_pairs = [
        ("Total records", d["total"]),
        ("Total active exceptions", d["total_active"]),
        ("HIGH risk (active)", d["high"]),
        ("MEDIUM risk (active)", d["medium"]),
        ("LOW risk (active)", d["low"]),
        ("Expiring this month", d["expiring_this_month"]),
        ("Due for renewal decision", d["expiring_due"]),
        ("Expired (not revoked)", d["expired_not_revoked"]),
        ("Overdue for review", d["overdue_for_review"]),
        ("Approvals recorded (%)", d["pct_approved"]),
        ("Admin/Root access (active)", d["by_type"]["admin_access"]),
        ("Firewall rules (active)", d["by_type"]["firewall_rule_open"]),
        ("Encryption waivers (active)", d["by_type"]["encryption_waiver"]),
    ]
    for label, value in summary_pairs:
        s.append([label, value])
    for i in range(4, 4 + len(summary_pairs)):
        s.cell(row=i, column=1).font = Font(color="334155")
        s.cell(row=i, column=2).font = Font(bold=True)
    _autosize(s, max_width=40)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _autosize(ws, max_width: int = 50) -> None:
    """Approximate column auto-fit based on cell content length."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)
